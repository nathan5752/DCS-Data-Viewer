"""
Data cleaning utilities for fixing common Excel file issues.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Tuple
import os
import config


class DataCleaner:
    """Cleans Excel files by fixing common data quality issues."""

    def __init__(self, filepath: str, tag_row: int,
                 description_row: int, units_row: int,
                 data_start_row: int):
        """
        Initialize cleaner with file parameters.

        Args:
            filepath: Path to the Excel file to clean
            tag_row: Row number containing tag names (1-indexed)
            description_row: Row number containing descriptions (1-indexed, can be None)
            units_row: Row number containing units (1-indexed, can be None)
            data_start_row: Row number where data starts (1-indexed)
        """
        self.filepath = filepath
        self.tag_row = tag_row
        self.description_row = description_row
        self.units_row = units_row
        self.data_start_row = data_start_row

    def clean_and_export(self) -> Tuple[bool, str, str]:
        """
        Clean the Excel file and export to a new file with '_cleaned' suffix.

        Returns:
            Tuple of (success, message, output_filepath)
        """
        try:
            # Generate output filename
            base_name = os.path.splitext(self.filepath)[0]
            extension = os.path.splitext(self.filepath)[1]
            output_filepath = f"{base_name}{config.CLEANED_FILE_SUFFIX}{extension}"

            # Step 1: Copy original file to output location (preserve original)
            import shutil
            shutil.copy2(self.filepath, output_filepath)

            # Step 2: Unmerge cells in the output file
            unmerge_success, unmerge_msg = self._unmerge_cells_in_file(output_filepath)

            # Step 3: Load the output file with pandas and perform cleaning
            df_full = pd.read_excel(output_filepath, header=None)

            # Step 4: Remove blank columns from the left
            df_full = self._remove_left_blank_columns(df_full)

            # Step 5: Fix duplicate column names in the tag row
            df_full = self._fix_duplicate_column_names(df_full)

            # Step 6: Remove rows with invalid timestamps
            df_full = self._remove_invalid_timestamp_rows(df_full)

            # Step 7: Export cleaned file (overwrite the copy we made)
            df_full.to_excel(output_filepath, index=False, header=False)

            message = f"Successfully cleaned file. Saved to: {os.path.basename(output_filepath)}"
            if unmerge_msg:
                message += f"\n{unmerge_msg}"

            return True, message, output_filepath

        except PermissionError:
            return False, f"Permission denied when writing to: {output_filepath}", ""
        except Exception as e:
            return False, f"Error cleaning file: {str(e)}", ""

    def _unmerge_cells_in_file(self, filepath: str) -> Tuple[bool, str]:
        """
        Unmerge cells in the specified Excel file.

        Args:
            filepath: Path to the Excel file to unmerge

        Returns:
            Tuple of (success, message)
        """
        try:
            workbook = load_workbook(filepath)
            sheet = workbook.active

            merged_ranges = list(sheet.merged_cells.ranges)

            if not merged_ranges:
                workbook.close()
                return True, ""

            unmerged_count = 0
            for merged_range in merged_ranges:
                # Get the top-left cell value
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                top_left_value = sheet.cell(min_row, min_col).value

                # Unmerge the range
                sheet.unmerge_cells(str(merged_range))

                # Fill all cells in the range with the top-left value
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        sheet.cell(row, col).value = top_left_value

                unmerged_count += 1

            # Save changes to the file
            workbook.save(filepath)
            workbook.close()

            return True, f"Unmerged {unmerged_count} cell range(s)"

        except Exception as e:
            return False, f"Warning: Could not unmerge cells: {str(e)}"

    def _remove_left_blank_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove blank columns from the left side of the dataframe.

        Args:
            df: Full dataframe with all rows

        Returns:
            DataFrame with left blank columns removed
        """
        # Check the tag row for blank columns
        tag_row_data = df.iloc[self.tag_row - 1]

        # Find first non-blank column
        first_data_col = 0
        for idx, value in enumerate(tag_row_data):
            if pd.notna(value) and str(value).strip() != '':
                first_data_col = idx
                break

        # Remove columns before first data column
        if first_data_col > 0:
            df = df.iloc[:, first_data_col:]

        return df

    def _fix_duplicate_column_names(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Fix duplicate column names by adding numeric suffixes.

        Args:
            df: Full dataframe with all rows

        Returns:
            DataFrame with fixed column names
        """
        # Get the tag row
        tag_row_idx = self.tag_row - 1
        tag_row_data = df.iloc[tag_row_idx].copy()

        # Track seen names and add suffixes to duplicates
        seen = {}
        new_names = []

        for idx, col_name in enumerate(tag_row_data):
            col_str = str(col_name) if pd.notna(col_name) else f"Column_{idx}"

            if col_str in seen:
                # Duplicate found - add suffix
                seen[col_str] += 1
                new_name = f"{col_str}_{seen[col_str]}"
                new_names.append(new_name)
            else:
                # First occurrence
                seen[col_str] = 0
                new_names.append(col_str)

        # Update the tag row
        # Convert entire DataFrame to object dtype to prevent FutureWarning
        # This is necessary because columns may have numeric dtypes (float64, int64)
        # and assigning string values across those columns triggers dtype incompatibility
        df = df.astype('object')
        df.iloc[tag_row_idx] = new_names

        return df

    def _remove_invalid_timestamp_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove rows with invalid timestamps from the data region.

        Args:
            df: Full dataframe with all rows

        Returns:
            DataFrame with invalid timestamp rows removed
        """
        # Get the data rows (starting from data_start_row)
        data_start_idx = self.data_start_row - 1

        # Separate header rows and data rows
        header_df = df.iloc[:data_start_idx]
        data_df = df.iloc[data_start_idx:]

        if len(data_df) == 0:
            return df

        # Try to parse first column as timestamps
        first_col = data_df.iloc[:, 0]
        parsed = pd.to_datetime(first_col, errors='coerce', format='ISO8601')

        # Keep only rows with valid timestamps
        valid_mask = parsed.notna()
        data_df_cleaned = data_df[valid_mask]

        # Recombine header and cleaned data
        result_df = pd.concat([header_df, data_df_cleaned], ignore_index=True)

        return result_df
