"""
Data validation module for checking Excel file quality before import.
"""

import pandas as pd
import numpy as np
from typing import Optional, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
import config


class ValidationIssue:
    """Represents a single data quality issue."""

    def __init__(self, severity: str, category: str, message: str,
                 location: Optional[str] = None, suggestion: Optional[str] = None):
        """
        Initialize a validation issue.

        Args:
            severity: Issue severity level ("ERROR", "WARNING", "INFO")
            category: Issue category (e.g., "Structure", "Data Type")
            message: Description of the issue
            location: Location reference (e.g., "B1:C1", "Column TEMP_01")
            suggestion: Suggested fix for the issue
        """
        self.severity = severity
        self.category = category
        self.message = message
        self.location = location
        self.suggestion = suggestion

    def __repr__(self):
        loc = f" at {self.location}" if self.location else ""
        return f"[{self.severity}] {self.category}: {self.message}{loc}"


class ValidationReport:
    """Contains all validation issues found in a file."""

    def __init__(self):
        self.issues: List[ValidationIssue] = []

    def add_issue(self, severity: str, category: str, message: str,
                  location: Optional[str] = None, suggestion: Optional[str] = None):
        """Add a new issue to the report."""
        issue = ValidationIssue(severity, category, message, location, suggestion)
        self.issues.append(issue)

    def has_errors(self) -> bool:
        """Check if report contains any errors."""
        return any(issue.severity == "ERROR" for issue in self.issues)

    def has_warnings(self) -> bool:
        """Check if report contains any warnings."""
        return any(issue.severity == "WARNING" for issue in self.issues)

    def has_issues(self) -> bool:
        """Check if report contains any issues at all."""
        return len(self.issues) > 0

    def get_errors(self) -> List[ValidationIssue]:
        """Get all error-level issues."""
        return [i for i in self.issues if i.severity == "ERROR"]

    def get_warnings(self) -> List[ValidationIssue]:
        """Get all warning-level issues."""
        return [i for i in self.issues if i.severity == "WARNING"]

    def get_infos(self) -> List[ValidationIssue]:
        """Get all info-level issues."""
        return [i for i in self.issues if i.severity == "INFO"]

    def get_issue_count(self) -> dict:
        """Get count of issues by severity."""
        return {
            "errors": len(self.get_errors()),
            "warnings": len(self.get_warnings()),
            "infos": len(self.get_infos())
        }


class DataValidator:
    """Validates Excel file data quality before import."""

    def __init__(self, filepath: str, tag_row: int,
                 description_row: Optional[int], units_row: Optional[int],
                 data_start_row: int):
        """
        Initialize validator with file parameters.

        Args:
            filepath: Path to the Excel file
            tag_row: Row number containing tag names (1-indexed)
            description_row: Row number containing descriptions (1-indexed, optional)
            units_row: Row number containing units (1-indexed, optional)
            data_start_row: Row number where data starts (1-indexed)
        """
        self.filepath = filepath
        self.tag_row = tag_row
        self.description_row = description_row
        self.units_row = units_row
        self.data_start_row = data_start_row

    def run_quick_validation(self) -> ValidationReport:
        """
        Run lightweight validation checks for automatic validation during load.

        Returns:
            ValidationReport with quick check results
        """
        report = ValidationReport()

        try:
            # Quick checks - prioritize speed over thoroughness
            self._check_merged_cells(report, quick=True)
            self._check_blank_columns(report)
            self._check_duplicate_column_names(report)
            self._check_timestamp_format(report, quick=True)
            self._check_numeric_data(report, quick=True)
        except Exception as e:
            report.add_issue(
                "ERROR",
                "Validation",
                f"Validation failed: {str(e)}",
                suggestion="Check that the file is a valid Excel file and not corrupted"
            )

        return report

    def run_full_validation(self) -> ValidationReport:
        """
        Run comprehensive validation checks for manual validation.

        Returns:
            ValidationReport with full check results
        """
        report = ValidationReport()

        try:
            # All checks - thorough analysis
            self._check_merged_cells(report, quick=False)
            self._check_blank_columns(report)
            self._check_duplicate_column_names(report)
            self._check_timestamp_format(report, quick=False)
            self._check_numeric_data(report, quick=False)
            self._check_missing_values(report)
            self._check_special_characters(report)
        except Exception as e:
            report.add_issue(
                "ERROR",
                "Validation",
                f"Validation failed: {str(e)}",
                suggestion="Check that the file is a valid Excel file and not corrupted"
            )

        return report

    def _check_merged_cells(self, report: ValidationReport, quick: bool = True):
        """Check for merged cells in the data region."""
        try:
            # Note: Cannot use read_only=True as it prevents access to merged_cells attribute
            workbook = load_workbook(self.filepath, data_only=True)
            sheet = workbook.active

            merged_ranges = list(sheet.merged_cells.ranges)

            if not merged_ranges:
                return  # No merged cells found

            # Check if merged cells are in the data region
            for merged_range in merged_ranges[:10] if quick else merged_ranges:
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                max_row = merged_range.max_row
                max_col = merged_range.max_col

                # Check if merged cell overlaps with our data region
                if min_row <= self.data_start_row or \
                   (min_row >= self.tag_row and min_row < self.data_start_row):

                    location = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
                    report.add_issue(
                        "ERROR",
                        "Structure",
                        f"Merged cells detected in data region",
                        location=location,
                        suggestion="Unmerge cells and fill with appropriate values"
                    )

                    if quick and len(report.get_errors()) >= 5:
                        break  # Limit errors in quick mode

            workbook.close()

        except Exception as e:
            report.add_issue(
                "WARNING",
                "Structure",
                f"Could not check for merged cells: {str(e)}"
            )

    def _check_blank_columns(self, report: ValidationReport):
        """Check for blank columns at the left."""
        try:
            # Read just the tag row to check for blank columns
            df = pd.read_excel(self.filepath, header=self.tag_row - 1, nrows=1)

            # Count blank columns from the left
            blank_count = 0
            for col in df.columns:
                if pd.isna(col) or str(col).strip() == '' or str(col).lower().startswith('unnamed'):
                    blank_count += 1
                else:
                    break

            if blank_count > 0:
                report.add_issue(
                    "WARNING",
                    "Structure",
                    f"Found {blank_count} blank column(s) at the left of the file",
                    suggestion="Remove blank columns or they will be skipped during import"
                )

        except Exception as e:
            report.add_issue(
                "WARNING",
                "Structure",
                f"Could not check for blank columns: {str(e)}"
            )

    def _check_duplicate_column_names(self, report: ValidationReport):
        """Check for duplicate tag names."""
        try:
            # Read the tag row
            df = pd.read_excel(self.filepath, header=self.tag_row - 1, nrows=0)

            # Get column names
            columns = [str(col) for col in df.columns if not (pd.isna(col) or str(col).lower().startswith('unnamed'))]

            # Find duplicates
            seen = {}
            duplicates = []
            for idx, col in enumerate(columns):
                if col in seen:
                    duplicates.append((col, seen[col], idx))
                else:
                    seen[col] = idx

            if duplicates:
                for col_name, first_idx, second_idx in duplicates:
                    report.add_issue(
                        "ERROR",
                        "Structure",
                        f"Duplicate column name '{col_name}' found",
                        location=f"Columns {get_column_letter(first_idx + 1)} and {get_column_letter(second_idx + 1)}",
                        suggestion="Rename duplicate columns to have unique names"
                    )

        except Exception as e:
            report.add_issue(
                "WARNING",
                "Structure",
                f"Could not check for duplicate columns: {str(e)}"
            )

    def _check_timestamp_format(self, report: ValidationReport, quick: bool = True):
        """Check if first column has valid timestamps."""
        try:
            # Read data starting from data_start_row
            skip_rows = list(range(0, self.tag_row - 1)) + list(range(self.tag_row, self.data_start_row - 1))

            # Read limited rows in quick mode
            nrows = config.VALIDATION_SAMPLE_ROWS if quick else None
            df = pd.read_excel(
                self.filepath,
                header=self.tag_row - 1,
                skiprows=skip_rows,
                nrows=nrows
            )

            if len(df) == 0:
                report.add_issue(
                    "ERROR",
                    "Data Type",
                    "No data rows found in the file",
                    suggestion="Ensure data starts at the specified data start row"
                )
                return

            # Check first column for timestamps
            first_col = df.columns[0]
            timestamp_col = df[first_col]

            # Try to parse as datetime
            parsed = pd.to_datetime(timestamp_col, errors='coerce', format='ISO8601')
            invalid_count = parsed.isna().sum()

            if invalid_count > 0:
                percentage = (invalid_count / len(timestamp_col)) * 100
                severity = "ERROR" if percentage > 10 else "WARNING"

                report.add_issue(
                    severity,
                    "Data Type",
                    f"{invalid_count} invalid timestamp(s) found ({percentage:.1f}% of data)",
                    location=f"Column {first_col}",
                    suggestion="Ensure first column contains valid date/time values"
                )

        except Exception as e:
            report.add_issue(
                "WARNING",
                "Data Type",
                f"Could not validate timestamps: {str(e)}"
            )

    def _check_numeric_data(self, report: ValidationReport, quick: bool = True):
        """Check if data columns contain numeric values."""
        try:
            # Read data
            skip_rows = list(range(0, self.tag_row - 1)) + list(range(self.tag_row, self.data_start_row - 1))
            nrows = config.VALIDATION_SAMPLE_ROWS if quick else None
            df = pd.read_excel(
                self.filepath,
                header=self.tag_row - 1,
                skiprows=skip_rows,
                nrows=nrows
            )

            if len(df) == 0:
                return

            # Check each column except the first (timestamp)
            data_columns = df.columns[1:]

            for col in data_columns:
                # Try to convert to numeric
                numeric_col = pd.to_numeric(df[col], errors='coerce')
                non_numeric_count = numeric_col.isna().sum() - df[col].isna().sum()

                if non_numeric_count > 0:
                    percentage = (non_numeric_count / len(df[col])) * 100
                    severity = "ERROR" if percentage > 10 else "WARNING"

                    report.add_issue(
                        severity,
                        "Data Type",
                        f"{non_numeric_count} non-numeric value(s) found in data column ({percentage:.1f}%)",
                        location=f"Column {col}",
                        suggestion="Ensure data columns contain only numeric values"
                    )

                # Limit checks in quick mode
                if quick and len([i for i in report.issues if i.category == "Data Type"]) >= 5:
                    break

        except Exception as e:
            report.add_issue(
                "WARNING",
                "Data Type",
                f"Could not validate numeric data: {str(e)}"
            )

    def _check_missing_values(self, report: ValidationReport):
        """Check for excessive missing values per column."""
        try:
            # Read all data
            skip_rows = list(range(0, self.tag_row - 1)) + list(range(self.tag_row, self.data_start_row - 1))
            df = pd.read_excel(
                self.filepath,
                header=self.tag_row - 1,
                skiprows=skip_rows
            )

            if len(df) == 0:
                return

            # Check each column except the first (timestamp)
            data_columns = df.columns[1:]

            for col in data_columns:
                missing_count = df[col].isna().sum()
                percentage = (missing_count / len(df[col])) * 100

                if percentage > config.VALIDATION_MAX_MISSING_PERCENT:
                    report.add_issue(
                        "WARNING",
                        "Completeness",
                        f"Column has {percentage:.1f}% missing values ({missing_count} of {len(df[col])})",
                        location=f"Column {col}",
                        suggestion="Consider removing this column or filling missing values"
                    )

        except Exception as e:
            report.add_issue(
                "INFO",
                "Completeness",
                f"Could not check missing values: {str(e)}"
            )

    def _check_special_characters(self, report: ValidationReport):
        """Check for special characters in tag names that might cause issues."""
        try:
            # Read the tag row
            df = pd.read_excel(self.filepath, header=self.tag_row - 1, nrows=0)

            # Check each column name
            problematic_chars = re.compile(r'[^\w\s\-_.]')

            for col in df.columns:
                col_str = str(col)
                if problematic_chars.search(col_str):
                    report.add_issue(
                        "INFO",
                        "Structure",
                        f"Column name contains special characters: '{col_str}'",
                        location=f"Column {col_str}",
                        suggestion="Consider using only letters, numbers, spaces, hyphens, underscores, and periods"
                    )

        except Exception as e:
            report.add_issue(
                "INFO",
                "Structure",
                f"Could not check for special characters: {str(e)}"
            )
